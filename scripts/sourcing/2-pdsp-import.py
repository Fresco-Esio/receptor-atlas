#!/usr/bin/env python3
"""Step 2 (replaces 2-pdsp-pull.mjs) — import the PDSP Ki Database from its XLSX export.

    python scripts/sourcing/2-pdsp-import.py KiDatabase_YYYY-MM-DD_HH-MM.xlsx

Writes cache/pdsp-rows.json (the shape 3-build.mjs consumes) plus cache/pdsp-snapshot.json
recording the file name, its SHA-256 and the row count, so a build is reproducible from a
named, checksummed input instead of from whatever the website served that afternoon.

Why this replaced the scraper
-----------------------------
The old step 2 paged an HTML results grid and parsed it with a regex over <td> cells. Two
things were wrong with that, and only the second was visible:

  1. The grid RENDERS A CENSORED SCREEN AS A BARE NUMBER. PDSP stores "we tested this pair
     and saw no meaningful binding" as the string ">10000"; the grid drops the ">". The
     scrape therefore recorded 1028 censored screens as measured Ki of exactly 10000, and
     nothing downstream could tell them apart. This file carries the ">" — and contains
     ZERO genuinely measured Ki of 10000, which is what proves the old reading was lossy.
  2. The database is live. Rows are added and deleted under you: a value present in the
     2026-07 scrape (Ki id 85425, fluoxetine at SERT) is already gone from this export.
     Verification against a moving target is not verification, hence the checksum.

Toolchain note: this step is Python because the export is XLSX and the repo has no Node
spreadsheet dependency. It runs once per refresh and its output is plain JSON, so the rest
of the pipeline stays Node.
"""
import hashlib
import json
import sys
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
CACHE = HERE / "cache"

# The export's column order. Asserted against the header row so a reshuffled export fails
# loudly here instead of silently mis-assigning species to ligand three steps downstream.
COLUMNS = ["Ki ID", "Receptor", "Sources", "Species", "Hot Ligands",
           "Test Ligands", "Ki Value", "Citations", "Submitted At", "Updated At"]

CENSORED = ">10000"          # PDSP's only qualifier form; see the module docstring


def main(path: Path) -> None:
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(min_row=1, values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)][:len(COLUMNS)]
    if header != COLUMNS:
        raise SystemExit(f"unexpected columns\n  expected {COLUMNS}\n  found    {header}")

    out, censored, unparsed = [], 0, 0
    for r in rows:
        if r[0] is None:
            continue
        value = r[6]
        is_censored = isinstance(value, str) and value.strip() == CENSORED
        if is_censored:
            ki, censored = None, censored + 1
        else:
            try:
                ki = float(value)
            except (TypeError, ValueError):
                # Any other non-numeric form is a qualifier this importer does not know.
                # Keep the row, flag it, and let the build decide — never coerce it.
                ki, unparsed = None, unparsed + 1
        out.append({
            "kiId": str(r[0]),
            "receptor": r[1] or "",
            "tissue": r[2] or "",
            "species": r[3] or "",
            "hot": r[4] or "",
            "test": r[5] or "",
            "ki": ki,
            "censored": is_censored,
            "raw": None if ki is not None else (str(value).strip() if value is not None else ""),
            "cite": r[7] or "",
            "updated": str(r[9] or ""),
        })
    wb.close()

    CACHE.mkdir(exist_ok=True)
    (CACHE / "pdsp-rows.json").write_text(json.dumps(out), encoding="utf8")
    (CACHE / "pdsp-snapshot.json").write_text(json.dumps({
        "file": path.name, "sha256": digest, "rows": len(out),
        "censored": censored, "unparsed": unparsed,
    }, indent=2), encoding="utf8")

    print(f"{path.name}\n  sha256    {digest}\n  rows      {len(out)}\n"
          f"  censored  {censored}\n  unparsed  {unparsed}")
    if unparsed:
        print("  ^ unrecognised qualifier forms — inspect before trusting the build")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        raise SystemExit(__doc__)
    main(Path(sys.argv[1]))
